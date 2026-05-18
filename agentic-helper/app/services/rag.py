from collections import defaultdict
from io import BytesIO
from uuid import uuid4

from langchain_text_splitters import RecursiveCharacterTextSplitter
from openpyxl import load_workbook

from app.core.config import settings
from app.models.knowledge import KnowledgeChunk
from app.schemas.rag import IngestDocument
from app.services.embeddings import EmbeddingService

try:
    from qdrant_client import models as qdrant_models
except ImportError:  # pragma: no cover - optional until dependencies are installed
    qdrant_models = None  # type: ignore[assignment]

_UNSET = object()


class RagService:
    def __init__(self, embeddings: EmbeddingService, qdrant_client) -> None:
        self._embeddings = embeddings
        self._qdrant_client = qdrant_client
        self._vector_name: str | None | object = _UNSET
        self._splitter = RecursiveCharacterTextSplitter(
            chunk_size=settings.rag_chunk_size,
            chunk_overlap=settings.rag_chunk_overlap,
        )

    @staticmethod
    def _build_source_filter(source: str):
        if qdrant_models is None:
            raise RuntimeError("qdrant-client is not installed")
        return qdrant_models.Filter(
            must=[
                qdrant_models.FieldCondition(
                    key="source",
                    match=qdrant_models.MatchValue(value=source),
                )
            ]
        )

    @staticmethod
    def _record_to_chunk(record) -> KnowledgeChunk:
        payload = record.payload or {}
        return KnowledgeChunk(
            id=str(record.id),
            source=str(payload.get("source", "")),
            content=str(payload.get("content", "")),
            metadata_json=payload.get("metadata_json") or {},
            score=getattr(record, "score", None),
        )

    async def _build_point_struct(
        self,
        *,
        point_id: str,
        source: str,
        content: str,
        metadata_json: dict,
    ):
        if qdrant_models is None:
            raise RuntimeError("qdrant-client is not installed")

        vector_name = await self._resolve_vector_name()
        embedding = await self._embeddings.embed_query(content)
        vector = {vector_name: embedding} if vector_name else embedding
        return qdrant_models.PointStruct(
            id=point_id,
            vector=vector,
            payload={
                "source": source,
                "content": content,
                "metadata_json": metadata_json,
            },
        )

    async def _resolve_vector_name(self) -> str | None:
        if self._vector_name is not _UNSET:
            return self._vector_name  # type: ignore[return-value]

        if settings.qdrant_vector_name.strip():
            self._vector_name = settings.qdrant_vector_name.strip()
            return self._vector_name

        collection = await self._qdrant_client.get_collection(settings.qdrant_collection_name)
        vectors = collection.config.params.vectors
        if isinstance(vectors, dict) and len(vectors) == 1:
            self._vector_name = next(iter(vectors))
        else:
            self._vector_name = None
        return self._vector_name

    async def ingest_documents(self, docs: list[IngestDocument], replace_source: bool = False) -> int:
        if qdrant_models is None:
            raise RuntimeError("qdrant-client is not installed")

        vector_name = await self._resolve_vector_name()

        if replace_source:
            by_source = defaultdict(list)
            for doc in docs:
                by_source[doc.source].append(doc)
            for source in by_source:
                await self._qdrant_client.delete(
                    collection_name=settings.qdrant_collection_name,
                    points_selector=self._build_source_filter(source),
                    wait=True,
                )

        pending: list = []
        chunks_created = 0

        for doc in docs:
            chunks = self._splitter.split_text(doc.text)
            if not chunks:
                continue
            vectors = await self._embeddings.embed_documents(chunks)
            for i, chunk in enumerate(chunks):
                vector = {vector_name: vectors[i]} if vector_name else vectors[i]
                pending.append(
                    qdrant_models.PointStruct(
                        id=str(uuid4()),
                        vector=vector,
                        payload={
                            "source": doc.source,
                            "content": chunk,
                            "metadata_json": {**doc.metadata, "chunk_index": i},
                        },
                    )
                )
            chunks_created += len(chunks)

        if pending:
            await self._qdrant_client.upsert(
                collection_name=settings.qdrant_collection_name,
                points=pending,
                wait=True,
            )
        return chunks_created

    async def retrieve(self, query: str) -> list[KnowledgeChunk]:
        query_vector = await self._embeddings.embed_query(query)
        vector_name = await self._resolve_vector_name()
        result = await self._qdrant_client.query_points(
            collection_name=settings.qdrant_collection_name,
            query=query_vector,
            using=vector_name,
            limit=settings.rag_top_k,
            with_payload=True,
            with_vectors=False,
        )
        points = getattr(result, "points", [])
        return [self._record_to_chunk(point) for point in points]

    async def list_chunks(
        self,
        *,
        limit: int = 20,
        offset: str | int | None = None,
        source: str | None = None,
    ) -> tuple[list[KnowledgeChunk], str | int | None]:
        records, next_offset = await self._qdrant_client.scroll(
            collection_name=settings.qdrant_collection_name,
            scroll_filter=self._build_source_filter(source) if source else None,
            limit=limit,
            offset=offset,
            with_payload=True,
            with_vectors=False,
        )
        return [self._record_to_chunk(record) for record in records], next_offset

    async def search_chunks(
        self,
        *,
        query: str,
        limit: int = 20,
        source: str | None = None,
    ) -> list[KnowledgeChunk]:
        query_vector = await self._embeddings.embed_query(query)
        vector_name = await self._resolve_vector_name()
        result = await self._qdrant_client.query_points(
            collection_name=settings.qdrant_collection_name,
            query=query_vector,
            using=vector_name,
            query_filter=self._build_source_filter(source) if source else None,
            limit=limit,
            with_payload=True,
            with_vectors=False,
        )
        return [self._record_to_chunk(point) for point in getattr(result, "points", [])]

    async def get_chunk(self, point_id: str) -> KnowledgeChunk | None:
        records = await self._qdrant_client.retrieve(
            collection_name=settings.qdrant_collection_name,
            ids=[point_id],
            with_payload=True,
            with_vectors=False,
        )
        if not records:
            return None
        return self._record_to_chunk(records[0])

    async def create_chunk(
        self,
        *,
        source: str,
        content: str,
        metadata_json: dict | None = None,
    ) -> KnowledgeChunk:
        point_id = str(uuid4())
        point = await self._build_point_struct(
            point_id=point_id,
            source=source,
            content=content,
            metadata_json=metadata_json or {},
        )
        await self._qdrant_client.upsert(
            collection_name=settings.qdrant_collection_name,
            points=[point],
            wait=True,
        )
        return KnowledgeChunk(
            id=point_id,
            source=source,
            content=content,
            metadata_json=metadata_json or {},
        )

    async def update_chunk(
        self,
        point_id: str,
        *,
        source: str | None = None,
        content: str | None = None,
        metadata_json: dict | None = None,
    ) -> KnowledgeChunk | None:
        existing = await self.get_chunk(point_id)
        if existing is None:
            return None

        next_source = source if source is not None else existing.source
        next_content = content if content is not None else existing.content
        next_metadata = metadata_json if metadata_json is not None else existing.metadata_json

        point = await self._build_point_struct(
            point_id=point_id,
            source=next_source,
            content=next_content,
            metadata_json=next_metadata,
        )
        await self._qdrant_client.upsert(
            collection_name=settings.qdrant_collection_name,
            points=[point],
            wait=True,
        )
        return KnowledgeChunk(
            id=point_id,
            source=next_source,
            content=next_content,
            metadata_json=next_metadata,
        )

    async def delete_chunk(self, point_id: str) -> bool:
        existing = await self.get_chunk(point_id)
        if existing is None:
            return False
        await self._qdrant_client.delete(
            collection_name=settings.qdrant_collection_name,
            points_selector=[point_id],
            wait=True,
        )
        return True

    def documents_from_excel(
        self,
        file_bytes: bytes,
        source: str,
        question_column: str = "question",
        answer_column: str = "answer",
        sheet_name: str | None = None,
    ) -> list[IngestDocument]:
        wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = ws.iter_rows(values_only=True)
        headers_row = next(rows, None)
        if headers_row is None:
            return []

        header_map: dict[str, int] = {}
        for i, value in enumerate(headers_row):
            if value is None:
                continue
            name = str(value).strip().lower()
            if name:
                header_map[name] = i

        q_col = question_column.strip().lower()
        a_col = answer_column.strip().lower()
        if q_col not in header_map or a_col not in header_map:
            return []

        q_idx = header_map[q_col]
        a_idx = header_map[a_col]

        docs: list[IngestDocument] = []
        for row_number, row in enumerate(rows, start=2):
            question = str(row[q_idx]).strip() if q_idx < len(row) and row[q_idx] is not None else ""
            answer = str(row[a_idx]).strip() if a_idx < len(row) and row[a_idx] is not None else ""
            if not question and not answer:
                continue
            text = f"Question: {question}\nAnswer: {answer}"
            docs.append(
                IngestDocument(
                    source=source,
                    text=text,
                    metadata={"row_number": row_number, "question": question},
                )
            )
        return docs
